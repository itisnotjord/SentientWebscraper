from transformers import pipeline
import pandas as pd
from openpyxl import load_workbook

# Load the Excel workbook
wb = load_workbook("/app/Articles.xlsx")

# Initialize the sentiment analysis model
specific_model = pipeline("text-classification", model="ProsusAI/finbert")

# Function to perform sentiment analysis on a list of texts
def perform_sentiment_analysis(texts):
    texts = [text[:512] for text in texts]  # Limit text to 512 characters for the model
    results = specific_model(texts)
    return [f"{result['label'].capitalize()}, Score: {result['score']:.2f}" for result in results]

# Iterate over each sheet in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Read the data from the sheet into a DataFrame
    data = ws.values
    columns = next(data)[0:]  # Get the first row as column names
    df = pd.DataFrame(data, columns=columns)

    # Perform sentiment analysis on the articles
    articles = df["Text"].tolist()  # Assuming the article text is in the "Text" column
    sentiment_results = perform_sentiment_analysis(articles)

    # Add a new header for the sentiment results
    ws.cell(row=1, column=5, value="Sentiment")

    # Append the sentiment results to the new column
    for idx, result in enumerate(sentiment_results):
        ws.cell(row=idx + 2, column=5, value=result)

# Save the updated workbook
wb.save("/app/Articles.xlsx")
print("Sentiment analysis results appended to all sheets in the Excel file.")
